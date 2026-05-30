#!/usr/bin/env python3
"""Generate 10 synthetic XLSX files for the per-type fixture corpus.

ADR-028 measurement prereq. Uses openpyxl (the [xlsx] extra). 5 small
reference sheets (<50 rows) and 5 large tabular sheets (100+ rows).

The ``project-falcon-quarterly-revenue.xlsx`` file is the canary partner
for boundary-spanning row queries — it has Q1 and Q2 revenue rows that
must both surface for "compare Q1 and Q2 revenue" to score.

Run from repo root:
    python3 scripts/reflib/generate_xlsx_fixtures.py

Output: reference-library/per-type-fixtures/xlsx/*.xlsx (10 files)
"""

from __future__ import annotations

import random
import sys
from pathlib import Path

from openpyxl import Workbook

_SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(_SCRIPT_DIR.parent.parent))

from scripts.reflib._fixture_vocab import (  # noqa: E402 — sys.path tweak above
    AGENTS,
    PROJECTS,
    QUARTERLY_REVENUE,
    QUARTERS,
    TOPICS,
)

SEED = 28028
N_FIXTURES = 10


def _small_reference_sheet(wb: Workbook, project: str, _rng: random.Random) -> None:
    """Small reference sheet (<50 rows) — attendee roster style."""
    ws = wb.active
    ws.title = "roster"
    ws.append(["agent", "role", "project", "status"])
    for i, agent in enumerate(AGENTS):
        role = ["lead", "reviewer", "operator", "on-call", "documentation", "release"][i % 6]
        status = ["active", "active", "active", "rotating", "active", "on-leave"][i % 6]
        ws.append([agent, role, project, status])


def _large_tabular_sheet(wb: Workbook, project: str, rng: random.Random) -> None:
    """Large tabular sheet (100+ rows) — backlog ledger style."""
    ws = wb.active
    ws.title = "backlog"
    ws.append(["id", "agent", "topic", "quarter", "status", "estimate_days"])
    for i in range(120):
        ws.append(
            [
                f"BL-{i + 1:04d}",
                rng.choice(AGENTS),
                rng.choice(TOPICS),
                rng.choice(QUARTERS),
                rng.choice(["todo", "in-progress", "review", "done"]),
                rng.randint(1, 13),
            ]
        )


def _build_canary_revenue_sheet(output: Path) -> None:
    """Canary-target XLSX — Q1 + Q2 revenue rows must both surface."""
    wb = Workbook()
    ws = wb.active
    ws.title = "revenue"
    ws.append(["project", "quarter", "fiscal_year", "revenue_usd"])
    for project, by_quarter in QUARTERLY_REVENUE.items():
        for quarter, revenue in by_quarter.items():
            ws.append([project, quarter, "FY26", revenue])
    wb.save(str(output))


def generate(output_dir: Path) -> int:
    output_dir.mkdir(parents=True, exist_ok=True)
    rng = random.Random(SEED)
    written = 0
    # Canary sheet first
    _build_canary_revenue_sheet(output_dir / "project-falcon-quarterly-revenue.xlsx")
    written += 1
    # 4 more small reference sheets
    for i in range(4):
        project, _focus = PROJECTS[i % len(PROJECTS)]
        path = output_dir / f"{project}-roster-fy26.xlsx"
        wb = Workbook()
        _small_reference_sheet(wb, project, rng)
        wb.save(str(path))
        written += 1
    # 5 large tabular sheets
    for i in range(5):
        project, _focus = PROJECTS[i % len(PROJECTS)]
        path = output_dir / f"{project}-backlog-ledger-{i + 1:02d}.xlsx"
        wb = Workbook()
        _large_tabular_sheet(wb, project, rng)
        wb.save(str(path))
        written += 1
    return written


def main() -> int:
    repo_root = Path(__file__).resolve().parents[2]
    output_dir = repo_root / "reference-library" / "per-type-fixtures" / "xlsx"
    n = generate(output_dir)
    print(f"xlsx fixtures: wrote {n} files to {output_dir}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
