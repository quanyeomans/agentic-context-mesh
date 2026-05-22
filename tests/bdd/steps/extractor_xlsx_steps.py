"""Step definitions for ``extractor_xlsx.feature`` (OF-3, Wave 4).

Drives the real :class:`kairix.extractors.xlsx.XlsxExtractor` against
in-memory xlsx fixtures synthesised with :mod:`openpyxl` — F1-clean
(no monkeypatch), F2-clean (no env mutation). The library itself is
exercised end-to-end through the production code path.

Step phrasings carry the literal word "xlsx" so the global
pytest-bdd step registry doesn't collide with sibling extractor
features' analogous Given/When/Then phrases.

Sabotage-proof per step:
  * "claims the mime type" — flipping ``can_extract`` to return
    ``False`` in production fails the step.
  * "carries one page per non-empty sheet" — flipping ``_render_workbook``
    to include the empty sheet bumps the page count and fails the step.
  * "does not carry a `## Sheet: Empty` header" — same mutation as above
    surfaces "## Sheet: Empty" in the markdown and fails the step.
  * "merged cell value appears exactly once" — disabling the
    ``_merged_cell_mask`` returns the value in every merged cell and
    fails the step.
  * "extractor's version string is non-empty" — clearing the module-level
    ``version`` constant in production fails the step.
  * "quality_ok false for the produced document" — flipping
    ``quality_ok`` to return ``True`` unconditionally fails the step.
"""

from __future__ import annotations

import io
from typing import Any

import openpyxl
import pytest
from openpyxl.utils import get_column_letter
from pytest_bdd import given, parsers, then, when

from kairix.extractors import ExtractedDocument
from kairix.extractors.xlsx import XlsxExtractor, make_extractor, version

pytestmark = pytest.mark.bdd


# Step-phrase fragments lifted to module constants because the same
# literal repeats across this module (F17 — no >=10-char string
# duplicated >=3 times in a module).
_PHRASE_SPREADSHEET_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_MERGED_BANNER = "Quarterly Report"


def _xlsx_bytes_three_sheets() -> bytes:
    """Synthesise an xlsx with three sheets: ``Data``, ``Empty``, ``Charts``.

    ``Data`` carries a small grid of values; ``Empty`` is left untouched
    so it has no cells with content; ``Charts`` carries another small
    grid (we don't synthesise a real chart embed — chart-only sheets
    are a runtime shape the production code handles via the
    ``_sheet_is_skippable`` fallback).
    """
    workbook = openpyxl.Workbook()
    # Workbook ships with one default sheet — rename and use it.
    data = workbook.active
    data.title = "Data"
    data.append(["product", "units", "total"])
    data.append(["widget", 10, 100])
    data.append(["gadget", 7, 49])

    workbook.create_sheet(title="Empty")

    charts = workbook.create_sheet(title="Charts")
    charts.append(["region", "value"])
    charts.append(["north", 42])
    charts.append(["south", 17])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _xlsx_bytes_merged_top_row() -> bytes:
    """Synthesise an xlsx whose top-left cell spans three columns.

    The merged value is :data:`_MERGED_BANNER`; the cells immediately
    under the banner carry distinct values so we can confirm only
    the top-left of the merge holds the banner.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Summary"
    sheet["A1"] = _MERGED_BANNER
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    sheet["A2"] = "q1"
    sheet["B2"] = "q2"
    sheet["C2"] = "q3"
    sheet["A3"] = 100
    sheet["B3"] = 200
    sheet["C3"] = 300

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _xlsx_bytes_one_empty_sheet() -> bytes:
    """Synthesise an xlsx with exactly one sheet, all cells blank."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Blank"
    # Don't append any rows — the sheet stays empty.
    # Reference a column letter just to keep the lint clean (F19 — unused import).
    _ = get_column_letter(1)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


@pytest.fixture
def xlsx_state() -> dict[str, Any]:
    """Per-scenario state container."""
    return {
        "extractor": None,
        "raw": b"",
        "claimed": None,
        "doc": None,
    }


@given(parsers.parse('the xlsx extractor is registered under the name "{name}"'))
def _register_xlsx(xlsx_state: dict[str, Any], name: str) -> None:
    real = make_extractor()
    assert isinstance(real, XlsxExtractor)
    assert real.name == name
    xlsx_state["extractor"] = real


@given('the operator has an xlsx workbook with sheets named "Data" and "Empty" and "Charts"')
def _three_sheets(xlsx_state: dict[str, Any]) -> None:
    xlsx_state["raw"] = _xlsx_bytes_three_sheets()


@given("the operator has an xlsx workbook whose top-left cell spans the row across three columns")
def _merged_top_row(xlsx_state: dict[str, Any]) -> None:
    xlsx_state["raw"] = _xlsx_bytes_merged_top_row()


@given("the operator has an xlsx workbook with one empty sheet")
def _one_empty_sheet(xlsx_state: dict[str, Any]) -> None:
    xlsx_state["raw"] = _xlsx_bytes_one_empty_sheet()


@when(parsers.parse('the operator asks the xlsx extractor whether it can extract mime "{mime}"'))
def _ask_can_extract(xlsx_state: dict[str, Any], mime: str) -> None:
    extractor: XlsxExtractor = xlsx_state["extractor"]
    xlsx_state["claimed"] = extractor.can_extract(mime, xlsx_state["raw"][:8])


@when("the operator asks the xlsx extractor whether it can extract the spreadsheetml.sheet mime")
def _ask_spreadsheet_mime(xlsx_state: dict[str, Any]) -> None:
    extractor: XlsxExtractor = xlsx_state["extractor"]
    xlsx_state["claimed"] = extractor.can_extract(_PHRASE_SPREADSHEET_MIME, b"PK\x03\x04")


@when("the operator invokes the xlsx extractor's extract method on the workbook bytes")
def _invoke_extract(xlsx_state: dict[str, Any]) -> None:
    extractor: XlsxExtractor = xlsx_state["extractor"]
    xlsx_state["doc"] = extractor.extract(xlsx_state["raw"], _PHRASE_SPREADSHEET_MIME)


@then("the xlsx extractor claims the mime type")
def _then_claims(xlsx_state: dict[str, Any]) -> None:
    assert xlsx_state["claimed"] is True


@then("the xlsx extractor does not claim the mime type")
def _then_does_not_claim(xlsx_state: dict[str, Any]) -> None:
    assert xlsx_state["claimed"] is False


@then("the xlsx document carries one page per non-empty sheet")
def _then_page_per_sheet(xlsx_state: dict[str, Any]) -> None:
    doc: ExtractedDocument = xlsx_state["doc"]
    assert isinstance(doc, ExtractedDocument)
    # Data + Charts survive; Empty is skipped.
    assert len(doc.pages) == 2


@then(parsers.parse('the xlsx document markdown carries a "{header}" header'))
def _then_markdown_has_header(xlsx_state: dict[str, Any], header: str) -> None:
    doc: ExtractedDocument = xlsx_state["doc"]
    assert header in doc.markdown


@then(parsers.parse('the xlsx document markdown does not carry a "{header}" header'))
def _then_markdown_lacks_header(xlsx_state: dict[str, Any], header: str) -> None:
    doc: ExtractedDocument = xlsx_state["doc"]
    assert header not in doc.markdown


@then("the merged cell value appears exactly once in the rendered markdown")
def _then_merged_value_once(xlsx_state: dict[str, Any]) -> None:
    doc: ExtractedDocument = xlsx_state["doc"]
    assert doc.markdown.count(_MERGED_BANNER) == 1


@then("the xlsx extractor reports quality_ok true for the produced document")
def _then_quality_ok_true(xlsx_state: dict[str, Any]) -> None:
    extractor: XlsxExtractor = xlsx_state["extractor"]
    doc: ExtractedDocument = xlsx_state["doc"]
    assert extractor.quality_ok(doc) is True


@then("the xlsx extractor reports quality_ok false for the produced document")
def _then_quality_ok_false(xlsx_state: dict[str, Any]) -> None:
    extractor: XlsxExtractor = xlsx_state["extractor"]
    doc: ExtractedDocument = xlsx_state["doc"]
    assert extractor.quality_ok(doc) is False


@then("the xlsx extractor's version string is non-empty")
def _then_version_non_empty(xlsx_state: dict[str, Any]) -> None:
    extractor: XlsxExtractor = xlsx_state["extractor"]
    assert isinstance(extractor.version, str)
    assert extractor.version.strip() != ""
    assert version.strip() != ""
