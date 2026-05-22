"""Contract test for the ``xlsx`` extractor plugin (F43).

Imports the canonical fake AND the real implementation, then runs the
same :class:`Extractor` Protocol assertions against both. The fake
proves the test seam is real; the real impl proves the production
class satisfies the same shape — driven against a real openpyxl-built
workbook (synthesised in-memory) without monkeypatching the upstream
library (F1-clean).

Sabotage-proofs:

  * Deleting ``version`` from :mod:`kairix.extractors.xlsx`
    breaks ``test_extractor_declares_version``.
  * Flipping ``can_extract`` to ``return True`` for ``text/plain``
    on the real impl breaks ``test_real_rejects_plain_text``.
  * Flipping the quality gate's char threshold to ``0`` breaks
    ``test_quality_ok_false_on_empty_workbook``.
  * Disabling :func:`_render_workbook`'s empty-sheet skip breaks
    ``test_real_skips_empty_sheets``.
"""

from __future__ import annotations

import io
from collections.abc import Callable

import openpyxl
import pytest

from kairix.extractors import ExtractedDocument, Extractor
from kairix.extractors.xlsx import (
    XlsxExtractor,
)
from kairix.extractors.xlsx import (
    make_extractor as make_real_extractor,
)
from kairix.extractors.xlsx import (
    version as xlsx_version,
)
from tests.fakes import FakeXlsxExtractor

pytestmark = pytest.mark.contract


# Test fixture bytes — synthesised in-memory by openpyxl. Mirroring the
# spec's "sample.xlsx" expectation (3 sheets, Data + Empty + Charts;
# only Data + Charts survive the empty-sheet skip).
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_three_sheet_workbook_bytes() -> bytes:
    """Build an xlsx with three sheets: Data, Empty, Charts.

    Data + Charts carry small grids; Empty is left untouched. The
    production extractor skips Empty; both fakes' and real's
    contract assertions check the surviving page count.
    """
    workbook = openpyxl.Workbook()
    data = workbook.active
    data.title = "Data"
    data.append(["product", "units"])
    data.append(["widget", 10])
    workbook.create_sheet(title="Empty")
    charts = workbook.create_sheet(title="Charts")
    charts.append(["region", "value"])
    charts.append(["north", 42])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_one_empty_sheet_bytes() -> bytes:
    """Build an xlsx with one sheet, all cells blank."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Blank"
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


_THREE_SHEET_BYTES = _build_three_sheet_workbook_bytes()


_Factory = Callable[[], Extractor]


@pytest.fixture(
    params=[
        pytest.param(lambda: FakeXlsxExtractor(), id="fake"),
        pytest.param(lambda: make_real_extractor(), id="real"),
    ]
)
def _extractor(request: pytest.FixtureRequest) -> Extractor:
    factory: _Factory = request.param
    return factory()


@pytest.mark.contract
def test_xlsx_extractor_satisfies_protocol() -> None:
    """The real factory returns an instance that is a runtime ``Extractor``."""
    real = make_real_extractor()
    assert isinstance(real, Extractor)
    assert isinstance(real, XlsxExtractor)


@pytest.mark.contract
def test_extractor_declares_version() -> None:
    """F40 requirement — module-level ``version`` is non-empty."""
    assert isinstance(xlsx_version, str)
    assert xlsx_version.strip() != ""


@pytest.mark.contract
def test_real_factory_returns_xlsx_instance() -> None:
    """``make_extractor`` returns a real :class:`XlsxExtractor`."""
    real = make_real_extractor()
    assert isinstance(real, XlsxExtractor)
    assert real.name == "xlsx"


@pytest.mark.contract
def test_can_extract_claims_xlsx_mime(_extractor: Extractor) -> None:
    """Both fake and real claim the spreadsheetml.sheet mime."""
    assert _extractor.can_extract(_XLSX_MIME, b"PK\x03\x04") is True


@pytest.mark.contract
def test_can_extract_claims_sheet_suffix_by_magic(_extractor: Extractor) -> None:
    """Magic-byte + mime-ending-with-'sheet' catches provider-specific suffixes."""
    assert _extractor.can_extract("application/x-vnd-acme.sheet", b"PK\x03\x04") is True


@pytest.mark.contract
def test_real_rejects_plain_text() -> None:
    """The real impl refuses ``text/plain`` — that's passthrough's job."""
    real = make_real_extractor()
    assert real.can_extract("text/plain", b"hello") is False


@pytest.mark.contract
def test_real_rejects_bare_zip_without_sheet_mime() -> None:
    """ZIP magic alone is ambiguous; the real impl waits for a sheet mime."""
    real = make_real_extractor()
    assert real.can_extract("application/octet-stream", b"PK\x03\x04") is False


@pytest.mark.contract
def test_extract_returns_document_with_pages(_extractor: Extractor) -> None:
    """``extract`` produces an :class:`ExtractedDocument` with at least one Page."""
    doc = _extractor.extract(_THREE_SHEET_BYTES, _XLSX_MIME)
    assert isinstance(doc, ExtractedDocument)
    assert len(doc.pages) >= 1


@pytest.mark.contract
def test_quality_ok_true_on_substantive_workbook(_extractor: Extractor) -> None:
    """Quality gate passes when the workbook produces enough markdown."""
    doc = _extractor.extract(_THREE_SHEET_BYTES, _XLSX_MIME)
    assert _extractor.quality_ok(doc) is True


@pytest.mark.contract
def test_quality_ok_false_on_empty_workbook() -> None:
    """Quality gate fails when openpyxl recovers no sheet content."""
    real = make_real_extractor()
    doc = real.extract(_build_one_empty_sheet_bytes(), _XLSX_MIME)
    assert real.quality_ok(doc) is False


@pytest.mark.contract
def test_real_skips_empty_sheets() -> None:
    """Empty sheets contribute no Page — exactly two pages survive (Data + Charts)."""
    real = make_real_extractor()
    doc = real.extract(_THREE_SHEET_BYTES, _XLSX_MIME)
    assert len(doc.pages) == 2
    assert "## Sheet: Data" in doc.markdown
    assert "## Sheet: Charts" in doc.markdown
    assert "## Sheet: Empty" not in doc.markdown
