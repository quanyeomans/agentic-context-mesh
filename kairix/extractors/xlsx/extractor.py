"""openpyxl-backed xlsx extractor — sheet-as-document rendering.

Wraps :mod:`openpyxl` (MIT, commercial-safe) and adapts each worksheet
in an ``.xlsx`` file to a :class:`Page` in the returned
:class:`ExtractedDocument`. Tables render as pipe-syntax markdown;
merged cells collapse to the top-left value (other cells in the merge
render empty); formula cells resolve to their cached displayed value
(``data_only=True``). Empty sheets and chart-only sheets are skipped —
they would not contribute retrievable content and would dilute the
final markdown.

Spec ref: ``docs/architecture/connector-ingestion-architecture.md``
§10 (Wave 4 OF-3 — Office mixed-media).

Test seam: ``XlsxExtractor.__init__`` accepts a ``workbook_loader=``
callable for unit/contract tests; the default loader lazy-imports
:func:`openpyxl.load_workbook` so the upstream library is only
imported when ``extract()`` is actually called (F1-clean — no
monkeypatching).
"""

from __future__ import annotations

import tempfile
from collections.abc import Callable
from pathlib import Path
from typing import TYPE_CHECKING, Any, Protocol, cast

from kairix.extractors import (
    DocMetadata,
    ExtractedDocument,
    MimeType,
    Page,
)

if TYPE_CHECKING:
    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet


#: Canonical plugin name surfaced by the extractor registry.
PLUGIN_NAME = "xlsx"

#: Minimum total markdown length the plugin treats as "quality ok".
#: Anything shorter is treated as a parse failure (empty workbook /
#: corrupt file) and signals an escalation up the chain.
_QUALITY_MIN_CHARS = 100

#: The Office Open XML spreadsheet mime type (the only mime the plugin
#: claims natively).
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

#: ZIP-archive header — XLSX files wrap content in a ZIP (Office Open
#: XML). The full PK\x03\x04 leading bytes alone don't disambiguate
#: from DOCX / PPTX, so we additionally require the mime hint to end
#: with "sheet" before claiming the artefact.
_MAGIC_ZIP = b"PK\x03\x04"

#: Sentinel string for empty cell renderings in the markdown table.
#: openpyxl returns ``None`` for empty cells; the markdown column
#: separator needs a non-``None`` token.
_EMPTY_CELL = ""

#: Markdown header prefix for per-sheet sections. Lifted to a module
#: constant because the same string is referenced from multiple
#: helpers (F17 — no 10+ char string duplicated 3+ times).
_SHEET_HEADER_PREFIX = "## Sheet: "


class _WorkbookLoader(Protocol):
    """Wire-shape Protocol for the openpyxl ``load_workbook`` callable.

    openpyxl exposes ``load_workbook(filename, data_only=True, ...)``
    returning a :class:`openpyxl.workbook.workbook.Workbook`. We
    declare the Protocol here so a test can pass an in-memory fake
    without monkeypatching the upstream module (F1-clean).
    """

    def __call__(self, filename: Any, *, data_only: bool) -> Workbook:
        """Load an xlsx file into a workbook object."""


def _default_workbook_loader() -> _WorkbookLoader:
    """Lazy-import :func:`openpyxl.load_workbook`.

    openpyxl is declared as an *optional* dependency in
    ``pyproject.toml`` — operators not ingesting spreadsheets skip
    the install. Resolving the import inside the factory means
    ``import kairix.extractors.xlsx`` succeeds in environments
    without the upstream library; the ``ImportError`` only fires when
    ``make_extractor()`` is actually called.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover — import path validated by extract() test
        raise RuntimeError(
            "xlsx: the upstream 'openpyxl' package is not installed. "
            "fix: pip install 'Kairix-agentic-knowledge-mgt[xlsx]' "
            "to opt into the xlsx (sheet-as-document) extractor. "
            "next: re-run the connector sync; the xlsx extractor will then resolve."
        ) from exc
    return cast("_WorkbookLoader", load_workbook)


class XlsxExtractor:
    """:class:`Extractor` impl that delegates to :mod:`openpyxl`.

    The instance carries the :data:`version` declared in the package
    ``__init__`` so the value flows from one canonical declaration
    site (F40) through to ``documents_media.extractor_version`` on
    every produced document. Re-extraction sweeps trigger off a
    version diff per spec §5.6.

    Test seam: the constructor accepts ``workbook_loader=`` so a
    contract / unit test passes a synthetic loader without
    monkeypatching :mod:`openpyxl` (F1-clean).
    """

    def __init__(
        self,
        *,
        version: str,
        workbook_loader: Callable[[], _WorkbookLoader] = _default_workbook_loader,
    ) -> None:
        """Construct the extractor with explicit ``version`` + loader factory."""
        self.name: str = PLUGIN_NAME
        self.version: str = version
        self._workbook_loader = workbook_loader

    def can_extract(self, mime: MimeType, magic_bytes: bytes) -> bool:
        """``True`` for the xlsx mime, with a magic-byte + mime-hint fallback.

        The mime hint is the primary signal — operators routing a
        well-formed Content-Type header want the dispatch to match
        the declared format. If the mime hint ends with ``"sheet"``
        (e.g. an Office Open XML spreadsheet served with a
        provider-specific suffix) AND the magic bytes match the
        ZIP-archive header, the plugin claims the artefact.
        ``application/zip`` alone is NOT claimed — without the mime
        hint we can't tell XLSX from any other zip-wrapped file.
        """
        if isinstance(mime, str) and mime == _XLSX_MIME:
            return True
        if not magic_bytes.startswith(_MAGIC_ZIP):
            return False
        return bool(isinstance(mime, str) and mime.endswith("sheet"))

    def extract(self, raw: bytes, _mime: MimeType) -> ExtractedDocument:
        """Write ``raw`` to a tmp file and walk every worksheet.

        openpyxl's :func:`load_workbook` accepts a path or a binary
        stream; we use a path because the library's read-only mode
        is a touch more forgiving with mis-encoded blobs. ``data_only=True``
        makes formula cells return their cached displayed value
        rather than the formula text.

        The ``_mime`` argument is positional for Protocol compliance
        but unused — the registry only dispatches xlsx bytes to this
        plugin, so the mime hint adds no information here.

        Confidence is the text-length / byte-length ratio, capped
        at 1.0 — a heuristic for "did openpyxl recover content".
        The orchestrator consults :meth:`quality_ok` to decide on
        escalation; the float is surfaced for observability.
        """
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)
            tmp.write(raw)
        try:
            loader = self._workbook_loader()
            workbook: Workbook = loader(str(tmp_path), data_only=True)
            pages, sections = _render_workbook(workbook)
            metadata = _metadata_from(workbook)
        finally:
            try:
                tmp_path.unlink()
            except OSError:  # pragma: no cover — best-effort cleanup
                pass
        markdown = "\n\n".join(sections)
        confidence = _confidence_heuristic(markdown, raw)
        return ExtractedDocument(
            markdown=markdown,
            pages=tuple(pages),
            images=(),
            metadata=metadata,
            confidence=confidence,
        )

    def quality_ok(self, doc: ExtractedDocument) -> bool:
        """Escalation gate per spec §10 (Wave 4 OF-3).

        Returns ``True`` only when:

          * the produced document has at least one :class:`Page`, AND
          * the unified markdown has at least :data:`_QUALITY_MIN_CHARS`
            characters of content.

        An all-empty / corrupt workbook returns False so the
        orchestrator can route the artefact through the escalation
        chain (xlsx → pdf_fallback → ocr → vision, when wave 3+
        members are present).
        """
        if len(doc.pages) < 1:
            return False
        return len(doc.markdown) >= _QUALITY_MIN_CHARS


# ---------------------------------------------------------------------------
# Workbook → markdown helpers — extracted so each function stays under the
# F16 cognitive-complexity ceiling (15).
# ---------------------------------------------------------------------------


def _render_workbook(workbook: Workbook) -> tuple[list[Page], list[str]]:
    """Walk every worksheet; return (pages, markdown sections).

    Empty / chart-only sheets are skipped. Each surviving sheet
    becomes one :class:`Page` (numbered by 1-based sheet index) and
    one markdown section (header + table).
    """
    pages: list[Page] = []
    sections: list[str] = []
    for index, sheet in enumerate(workbook.worksheets, start=1):
        if _sheet_is_skippable(sheet):
            continue
        section = _render_sheet(sheet)
        pages.append(
            Page(
                page_number=index,
                text=section,
                has_images=_sheet_has_images(sheet),
            )
        )
        sections.append(section)
    return pages, sections


def _sheet_is_skippable(sheet: Worksheet) -> bool:
    """A sheet is skippable when it has no data rows.

    "Chart-only" sheets in openpyxl present as either a Chartsheet
    (no ``iter_rows``) or a worksheet whose ``max_row``/``max_column``
    is None or 0. Both shapes collapse to "no data rows" — which we
    treat as skippable. The chart embed itself doesn't ship as
    markdown; downstream OCR / vision waves handle chart content
    when they land.
    """
    if not hasattr(sheet, "iter_rows"):
        return True
    if not _sheet_has_any_value(sheet):
        return True
    return False


def _sheet_has_any_value(sheet: Worksheet) -> bool:
    """True if any cell in the sheet has a non-``None`` value."""
    for row in sheet.iter_rows(values_only=True):
        for cell in row:
            if cell is not None and str(cell).strip() != "":
                return True
    return False


def _sheet_has_images(sheet: Worksheet) -> bool:
    """True if the worksheet has any embedded images.

    openpyxl exposes ``Worksheet._images`` for embedded images.
    Returns False defensively when the attribute is missing.
    """
    images = getattr(sheet, "_images", None)
    if not images:
        return False
    return len(images) > 0


def _render_sheet(sheet: Worksheet) -> str:
    """Render one worksheet as a markdown section.

    Layout:

    .. code-block:: markdown

       ## Sheet: <title>

       | col1 | col2 | col3 |
       | --- | --- | --- |
       | a | b | c |
       | d |  | f |
    """
    header = f"{_SHEET_HEADER_PREFIX}{sheet.title}"
    merged_mask = _merged_cell_mask(sheet)
    rows = _collect_rows(sheet, merged_mask)
    if not rows:
        return header
    column_count = max(len(row) for row in rows)
    if column_count == 0:
        return header
    body = _format_table(rows, column_count)
    return f"{header}\n\n{body}"


def _merged_cell_mask(sheet: Worksheet) -> dict[tuple[int, int], bool]:
    """Map (row, col) -> True for cells in a merged range that are NOT
    the top-left of the merge. The top-left cell of every merge bears
    the value; the others render blank.

    Cell coordinates here are 1-based to match openpyxl's convention.
    """
    masked: dict[tuple[int, int], bool] = {}
    ranges = getattr(sheet, "merged_cells", None)
    if ranges is None:
        return masked
    for merge in list(ranges.ranges):
        top_row = merge.min_row
        top_col = merge.min_col
        for r in range(merge.min_row, merge.max_row + 1):
            for c in range(merge.min_col, merge.max_col + 1):
                if r == top_row and c == top_col:
                    continue
                masked[(r, c)] = True
    return masked


def _collect_rows(
    sheet: Worksheet,
    merged_mask: dict[tuple[int, int], bool],
) -> list[list[str]]:
    """Walk every row; render each cell as a string (empty for
    non-top-left merged cells). Drop trailing rows whose cells are
    all-empty — keeps the rendered table flush.
    """
    raw_rows: list[list[str]] = []
    for row_index, row in enumerate(sheet.iter_rows(), start=1):
        rendered: list[str] = []
        for col_index, cell in enumerate(row, start=1):
            if (row_index, col_index) in merged_mask:
                rendered.append(_EMPTY_CELL)
                continue
            rendered.append(_render_cell_value(cell.value))
        raw_rows.append(rendered)
    return _strip_trailing_empty_rows(raw_rows)


def _render_cell_value(value: Any) -> str:
    """Convert one cell's value to a markdown-safe string.

    ``None`` becomes the empty-cell sentinel. Pipes are escaped
    because they are the markdown table column separator. Newlines
    collapse to ``" "`` so a wrapped cell stays on one row.
    """
    if value is None:
        return _EMPTY_CELL
    text = str(value)
    text = text.replace("\n", " ").replace("\r", " ")
    text = text.replace("|", "\\|")
    return text.strip()


def _strip_trailing_empty_rows(rows: list[list[str]]) -> list[list[str]]:
    """Drop any trailing rows whose every cell is the empty-cell sentinel."""
    while rows and all(cell == _EMPTY_CELL for cell in rows[-1]):
        rows.pop()
    return rows


def _format_table(rows: list[list[str]], column_count: int) -> str:
    """Render the collected rows as a pipe-syntax markdown table.

    First row is the header; a separator row follows; remaining rows
    are the body. If only one row is present we still emit a separator
    so the result is a valid markdown table.
    """
    padded = [_pad_row(row, column_count) for row in rows]
    header_row = _format_row(padded[0])
    separator = "| " + " | ".join(["---"] * column_count) + " |"
    body_rows = [_format_row(row) for row in padded[1:]]
    parts = [header_row, separator, *body_rows]
    return "\n".join(parts)


def _pad_row(row: list[str], column_count: int) -> list[str]:
    """Pad a row out to ``column_count`` columns with empty cells."""
    if len(row) >= column_count:
        return row
    return row + [_EMPTY_CELL] * (column_count - len(row))


def _format_row(row: list[str]) -> str:
    """Render one table row as a pipe-syntax markdown line."""
    return "| " + " | ".join(row) + " |"


def _metadata_from(workbook: Workbook) -> DocMetadata:
    """Map ``workbook.properties`` to a :class:`DocMetadata` record.

    Empty / missing values stay as ``None``. ``created`` (a datetime)
    is rendered ISO-8601 when present.
    """
    props = getattr(workbook, "properties", None)
    title = _string_or_none(getattr(props, "title", None))
    author = _string_or_none(getattr(props, "creator", None))
    created = _datetime_or_none(getattr(props, "created", None))
    return DocMetadata(
        title=title,
        author=author,
        created_date=created,
        language=None,
        page_count=None,
    )


def _string_or_none(value: Any) -> str | None:
    """Return ``value`` as a string when it's a non-empty string, else ``None``."""
    if isinstance(value, str) and value.strip() != "":
        return value
    return None


def _datetime_or_none(value: Any) -> str | None:
    """Render a datetime as ISO-8601, else ``None``."""
    if value is None:
        return None
    iso = getattr(value, "isoformat", None)
    if callable(iso):
        result = iso()
        return result if isinstance(result, str) else None
    return None


def _confidence_heuristic(markdown: str, raw: bytes) -> float:
    """Cheap "did openpyxl actually recover content" signal.

    Returns the text-length / byte-length ratio, capped at 1.0.
    A blank workbook returns ~0.0; a dense text-only spreadsheet
    typically lands in the 0.1-0.4 range. The
    orchestrator consults :meth:`quality_ok` for the binary
    escalation decision; this float is for observability only.
    """
    if not raw:
        return 0.0
    return min(len(markdown) / len(raw), 1.0)
